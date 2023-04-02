from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
import optparse
import ipaddress
from openpyxl import load_workbook
import os
import select
import random
import string
import subprocess

sendEmailtoList = ['email@gmail.com','someoneelse@gmail.com']
sendEmailFrom = 'email@gmail.com'
smtpServerList = ['10.0.0.1','10.0.0.2', '10.0.0.3']

def email(subject, body):
	msg = MIMEMultipart()
	msg['Subject'] = subject
	msg['From'] = sendEmailFrom
	msg['To'] = ', '.join( sendEmailtoList )
	msg.preamble = 'MIME body below\n'
	msg.attach( MIMEText( body ))
	for s in smtpServerList:
		try:
			s = smtplib.SMTP( s ) 
			s.sendmail(msg['From'], sendEmailtoList, msg.as_string())
			s.close()
			break
		except:
			print('Warning:\tSMTP server %s failed, please remove the server from the server list' %s)
			continue
			
if __name__ == "__main__":
	parser = optparse.OptionParser(usage="Usage: %prog [OPTION]... <EXCEl_SPREADSHEET>...", description="Host discovery scan of each IPv4 network in a spreadsheet column of multiple spreadsheets.\nIP-address or full CIDR syntax is needed: x.x.x.x/y, only one per cell. Scan result is sent as per email once done. An notification email is sent if the timeout is reached.",
		epilog='Open Source MIT License. Written by Christian Angerbjorn')
	parser.add_option("-n", "--network", default='A', help="Network data column")
	parser.add_option("-t", "--timeout", default=60*60*24*2, type=int, help="Timeout in seconds")
	(ops, args) = parser.parse_args()

	if len(args) == 0:
		parser.error("At least one excel .xlsx file is required!")
		
	rand = ''.join(random.choice(string.ascii_uppercase + string.digits) for x in range(5))
	head = os.path.splitext( os.path.split(args[0])[1] )[0]	 
	proclog  =  head +'.'+ rand + '.log' # log for the process 
	hostfile =  head +'.'+ rand + '.hosts' # hosts to scan, fed to nmap 
	nmaplog  =  head +'.'+ rand + '.gnmap' # nmaps output 

	pl = open(proclog, 'wb')
	print( 'Created proc log: %s' %proclog )

	networks = [] 
	for excel in args:
		wb = load_workbook( filename = excel, read_only=True)
		for sheet in wb:
			for i, row in enumerate(sheet.iter_rows()):
				data = row[ ord(ops.network.upper())-65 ].value
				this_address = '%s:%s%d' %(excel, ops.network.upper(), i+1 )
				try:
					ipaddress.ip_network( data )
				except ValueError as err:
					if i and not data.startswith('#'): # first line assumed to be a header...
						print( "Warning: Cell %s%d value '%s' is not recognised as a IPv4 net." %(ops.network, i+1, data))
					continue
				networks.append( data )
	print( 'Created hostfile: "%s" with %d networks' %(hostfile, len(networks)))
	with open(hostfile, 'w') as f:
		f.write( '\n'.join( networks ) )
	
	# start sub process
	child = subprocess.Popen(['sudo', 'nmap', '-sn', '-n', '-PY', '-PE', '-PP', '-PS139,445,3389,443,80,8081,8080,5900,5061,6001,2381,623,23,22,161', '--host-timeout', '6m', '--max-retries', '1', '-T4', '-oG', nmaplog, '-iL', hostfile] , stdout=subprocess.PIPE, stderr=subprocess.STDOUT)

	# loop until dead
	alarm = False
	while child.poll() is None:
		if select.select([child.stdout],[],[], ops.timeout )[0] != []:
			pl.write(child.stdout.readline())
			pl.flush()
		else:
			pl.write(child.stdout.readline())
			pl.close()
			child.terminate()
			print('Host discovery failed: timeout reached for: %s' %proclog )
			email('Host discovery failed: timeout reached for: %s' %proclog, 'scan failed :(')
			exit()
	pl.write( child.communicate()[0] )
	child.stdout.close()
	# get IPS from output 
	awake = []
	with open(nmaplog, 'r') as f:
		for line in f.read().splitlines():
			if line.find('Status: Up') != -1:
				awake.append( line.split()[1] )
		print('Host discovery completed for: %s' %proclog)
		email('Host discovery completed for: %s' %proclog, 'The following hosts where discovered:\n' + '\n'.join( awake ))
	pl.close()
